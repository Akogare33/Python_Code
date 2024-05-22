import pandas as pd
import openpyxl
# 读取 Excel 数据到 DataFrame 对象
file_path = 'E:\\python脚本\\test.xlsx'
save_path = 'E:\\python脚本\\new.xlsx'
sheet_name='Sheet1'
df = pd.read_excel(file_path, sheet_name)




# 提取超链接数据中包含 "cloud3" 的值
Names = df['Name'].tolist()
Links = df['Link'].tolist()




# try:
for num in range(30):
    count = 0
    res = []
    word = f'CloudMac{num}_$'
    for name in Names:
        count += 1
        if word in name:
            # print(count)
            print(name[:-2])
            print(Links[count-1])

            res.append(name[:-2] + " : " + Links[count-1])
            # res.append(Links[count-1])
            

    wb = openpyxl.load_workbook(save_path)
    ws = wb[sheet_name]
    max_row_num = ws.max_row
    max_col_num = ws.max_column
    ws._current_row = max_row_num
    ws.append(res)
    wb.save(save_path)
# except Exception as e :
#     print(e)
#     pass

